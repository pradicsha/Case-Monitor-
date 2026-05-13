"""
Salesforce Critical & Urgent Case Monitor
------------------------------------------
A lightweight Flask dashboard that pulls Critical/Urgent Salesforce cases
created in the last 45 hours, flags them by SLA age (12h / 30h / 45h),
exports a styled Excel report, and (optionally) hands the report off to
a PowerShell uploader for SharePoint distribution.

Author : Nithin Gopal
"""

import os
import subprocess
import threading
import webbrowser
from datetime import datetime, timedelta, timezone
from pathlib import Path

import pandas as pd
from dotenv import load_dotenv
from flask import Flask, render_template
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from simple_salesforce import Salesforce

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
load_dotenv()

SF_USERNAME = os.getenv("SALESFORCE_USERNAME")
SF_PASSWORD = os.getenv("SALESFORCE_PASSWORD")
SF_TOKEN = os.getenv("SALESFORCE_SECURITY_TOKEN")
SF_INSTANCE = os.getenv("SALESFORCE_INSTANCE")

LOOKBACK_HOURS = int(os.getenv("LOOKBACK_HOURS", "45"))
REPORT_FILENAME = os.getenv("REPORT_FILENAME", "status_report.xlsx")
POWERSHELL_SCRIPT = os.getenv("POWERSHELL_SCRIPT_PATH", "").strip()

FLASK_HOST = os.getenv("FLASK_HOST", "127.0.0.1")
FLASK_PORT = int(os.getenv("FLASK_PORT", "5000"))
FLASK_DEBUG = os.getenv("FLASK_DEBUG", "true").lower() == "true"
AUTO_OPEN_BROWSER = os.getenv("AUTO_OPEN_BROWSER", "true").lower() == "true"

REQUIRED_VARS = {
    "SALESFORCE_USERNAME": SF_USERNAME,
    "SALESFORCE_PASSWORD": SF_PASSWORD,
    "SALESFORCE_SECURITY_TOKEN": SF_TOKEN,
    "SALESFORCE_INSTANCE": SF_INSTANCE,
}
missing = [name for name, val in REQUIRED_VARS.items() if not val]
if missing:
    raise RuntimeError(
        "Missing required environment variables: "
        + ", ".join(missing)
        + ". Copy .env.example to .env and fill in your Salesforce credentials."
    )

# ---------------------------------------------------------------------------
# Salesforce client
# ---------------------------------------------------------------------------
sf = Salesforce(
    username=SF_USERNAME,
    password=SF_PASSWORD,
    security_token=SF_TOKEN,
    instance_url=SF_INSTANCE,
)

# ---------------------------------------------------------------------------
# Flask app
# ---------------------------------------------------------------------------
app = Flask(__name__)


# ---------------------------------------------------------------------------
# Salesforce helpers
# ---------------------------------------------------------------------------
def _resolve_owner(owner_id, cache):
    """Look up an owner's Name and Email by Id, caching results to limit API calls."""
    if not owner_id:
        return {"Name": "Not Found", "Email": "Not Found"}
    if owner_id in cache:
        return cache[owner_id]

    user_query = f"SELECT Name, Email FROM User WHERE Id = '{owner_id}'"
    result = sf.query(user_query)
    records = result.get("records", [])
    if records:
        cache[owner_id] = {
            "Name": records[0].get("Name", "Not Found"),
            "Email": records[0].get("Email", "Not Found"),
        }
    else:
        cache[owner_id] = {"Name": "Not Found", "Email": "Not Found"}
    return cache[owner_id]


def fetch_open_critical_cases():
    """
    Pull open Critical/Urgent Cases created in the last LOOKBACK_HOURS.

    Returns a list of dicts ready for templating and Excel export.
    Each dict is enriched with:
        - Account            (flattened from Account.Name)
        - OwnerName, OwnerEmail
        - AgeHours           (float, hours since CreatedDate)
        - Alert_12h / Alert_30h / Alert_45h (SLA-window booleans)
    """
    try:
        now_utc = datetime.now(timezone.utc)
        lookback_iso = (now_utc - timedelta(hours=LOOKBACK_HOURS)).isoformat().replace("+00:00", "Z")
        print(f"[i] Fetching cases created since {lookback_iso}")

        soql = f"""
            SELECT Product_Line__c, Account.Name, CaseNumber, Case_Severity__c,
                   OwnerId, CreatedDate
            FROM Case
            WHERE Case_Severity__c IN ('Critical', 'Urgent')
              AND CreatedDate >= {lookback_iso}
              AND Status NOT IN ('Closed', 'Closed Pending')
            ORDER BY CreatedDate DESC
        """
        results = sf.query(soql)
        tickets = results.get("records", [])

        owner_cache = {}
        for ticket in tickets:
            # Strip Salesforce metadata
            ticket.pop("attributes", None)

            # Flatten Account.Name
            account = ticket.get("Account")
            ticket["Account"] = account.get("Name", "Unknown Account") if isinstance(account, dict) else "Unknown Account"

            # Compute case age and SLA flags
            created_str = ticket.get("CreatedDate")
            created_dt = datetime.strptime(created_str, "%Y-%m-%dT%H:%M:%S.%f%z")
            age_hours = (now_utc - created_dt).total_seconds() / 3600
            ticket["AgeHours"] = round(age_hours, 2)
            ticket["Alert_12h"] = 11 <= age_hours < 12
            ticket["Alert_30h"] = 29 <= age_hours < 30
            ticket["Alert_45h"] = 44 <= age_hours < 45

            # Resolve owner
            owner = _resolve_owner(ticket.get("OwnerId"), owner_cache)
            ticket["OwnerName"] = owner["Name"]
            ticket["OwnerEmail"] = owner["Email"]

        print(f"[+] Retrieved {len(tickets)} open critical/urgent case(s).")
        return tickets

    except Exception as exc:  # noqa: BLE001 - surface any SF/network failure
        print(f"[!] Error fetching tickets: {exc}")
        return []


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------
def export_to_excel(tickets, filename=REPORT_FILENAME):
    """Write the ticket list to a styled .xlsx report and return its absolute path."""
    if not tickets:
        print("[!] No tickets to export.")
        return None

    df = pd.DataFrame(tickets)
    if df.empty or not len(df.columns):
        print("[!] Empty DataFrame, skipping Excel generation.")
        return None

    filepath = Path(filename).resolve()
    df.to_excel(filepath, index=False)

    # Apply table styling
    wb = load_workbook(filepath)
    ws = wb.active
    last_col = get_column_letter(len(df.columns))
    table_range = f"A1:{last_col}{len(df) + 1}"

    table = Table(displayName="StatusReportTable", ref=table_range)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    wb.save(filepath)

    print(f"[+] Excel report saved: {filepath}")
    return str(filepath)


def upload_via_powershell(report_path):
    """
    Optional: run a PowerShell script to push the report to SharePoint.
    Skipped silently when POWERSHELL_SCRIPT_PATH is unset or the file is missing.
    """
    if not POWERSHELL_SCRIPT:
        print("[i] POWERSHELL_SCRIPT_PATH not set; skipping SharePoint upload.")
        return

    if not Path(POWERSHELL_SCRIPT).is_file():
        print(f"[!] PowerShell script not found at {POWERSHELL_SCRIPT}; skipping upload.")
        return

    try:
        subprocess.run(
            [
                "powershell.exe",
                "-ExecutionPolicy", "Bypass",
                "-File", POWERSHELL_SCRIPT,
                "-ReportPath", report_path or "",
            ],
            check=True,
        )
        print("[+] PowerShell upload script executed successfully.")
    except FileNotFoundError:
        print("[!] powershell.exe not available on this platform; skipping upload.")
    except subprocess.CalledProcessError as exc:
        print(f"[!] PowerShell script failed: {exc}")


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------
@app.route("/")
def index():
    tickets = fetch_open_critical_cases()
    report_path = export_to_excel(tickets)
    upload_via_powershell(report_path)
    return render_template(
        "index.html",
        tickets=tickets,
        generated_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        lookback_hours=LOOKBACK_HOURS,
    )


# ---------------------------------------------------------------------------
# Entrypoint
# ---------------------------------------------------------------------------
def _open_browser():
    webbrowser.open_new(f"http://{FLASK_HOST}:{FLASK_PORT}/")


if __name__ == "__main__":
    if AUTO_OPEN_BROWSER:
        threading.Timer(1.0, _open_browser).start()
    app.run(host=FLASK_HOST, port=FLASK_PORT, debug=FLASK_DEBUG)
